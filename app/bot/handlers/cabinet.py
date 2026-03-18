"""
/cabinet — кабинет администрирования.
Доступен только CRM-администраторам.
"""

from __future__ import annotations

import asyncio
import io
from datetime import datetime, timedelta, timezone

import openpyxl
from aiogram import F, Router
from aiogram.exceptions import TelegramBadRequest
from aiogram.filters import Command
from aiogram.types import BufferedInputFile, CallbackQuery, FSInputFile, InlineKeyboardButton, Message
from aiogram.utils.keyboard import InlineKeyboardBuilder
from loguru import logger
from openpyxl.styles import Alignment, Font, PatternFill

from app.bot.constants.ttl import TTL_ERROR_SEC, TTL_MENU_SEC
from app.bot.topic_cache import invalidate as invalidate_topic_cache
from app.bot.topic_resolver import resolve_topic_thread_id
from app.bot.topics import TOPIC_SPECS, TopicKey
from app.bot.utils.callback_parser import safe_parse
from app.bot.utils.handler_helpers import resolve_admin_context
from app.core.config import settings
from app.db.database import AsyncSessionLocal
from app.db.models.lead import LeadStatus
from app.db.repositories.lead_repository import LeadRepository
from app.db.repositories.tenant_topics import TenantTopicRepository
from app.telegram.html_utils import html_escape
from app.telegram.safe_sender import TelegramSafeSender

router = Router()

CABINET_HOME_TEXT = "🗂 <b>Кабинет</b>\n\nВыберите раздел."


def _get_topic_spec(key: TopicKey):
    for spec in TOPIC_SPECS:
        if spec.key == key:
            return spec
    return None


async def _probe_topic_thread(sender: TelegramSafeSender, chat_id: int, thread_id: int) -> bool:
    await sender.get_chat(chat_id)
    probe = await sender.send_text(chat_id=chat_id, message_thread_id=thread_id, text=".")
    try:
        await sender.delete_message(
            chat_id=chat_id,
            message_id=probe.message_id,
            thread_id=probe.message_thread_id,
        )
    except Exception as exc:
        logger.warning(f"Could not delete probe message in cabinet topic {thread_id}: {exc}")
    return True


async def _pin_cabinet_message(sender: TelegramSafeSender, chat_id: int, topic_id: int, message_id: int):
    try:
        await sender.unpin_all_forum_topic_messages(chat_id, topic_id)
    except Exception as exc:
        logger.warning(f"Could not unpin old cabinet messages: {exc}")
    try:
        await sender.pin_chat_message(chat_id, message_id, disable_notification=True)
    except Exception as exc:
        logger.warning(f"Could not pin cabinet message: {exc}")


async def _safe_edit_cabinet_message(
    sender: TelegramSafeSender,
    repo: LeadRepository,
    chat_id: int,
    topic_id: int,
    message_id: int,
    text: str,
    reply_markup,
) -> int:
    try:
        await sender.edit_message_text(
            chat_id=chat_id,
            message_id=message_id,
            text=text,
            reply_markup=reply_markup,
            parse_mode="HTML",
            thread_id=topic_id,
        )
        return message_id
    except TelegramBadRequest as e:
        msg = str(e).lower()
        if "message is not modified" in msg:
            try:
                await sender.edit_message_reply_markup(
                    chat_id=chat_id,
                    message_id=message_id,
                    reply_markup=reply_markup,
                    thread_id=topic_id,
                )
            except TelegramBadRequest as markup_err:
                logger.warning(f"Cabinet reply markup edit failed: {markup_err}")
            return message_id
        if any(token in msg for token in ("message to edit not found", "message_id_invalid", "message can't be edited")):
            new_msg = await sender.send_message(
                chat_id=chat_id,
                message_thread_id=topic_id,
                text=text,
                reply_markup=reply_markup,
                parse_mode="HTML",
            )
            await repo.set_panel_message_id(chat_id, topic_id, new_msg.message_id)
            await _pin_cabinet_message(sender, chat_id, topic_id, new_msg.message_id)
            logger.warning(f"Cabinet message restored with new message_id={new_msg.message_id}")
            return new_msg.message_id
        logger.error(f"Failed to edit cabinet message: {e}")
        return message_id


async def ensure_cabinet_message(sender: TelegramSafeSender, chat_id: int) -> tuple[int, int] | None:
    async with AsyncSessionLocal() as session:
        repo = LeadRepository(session)
        topic_repo = TenantTopicRepository(session)

        topic_id = await resolve_topic_thread_id(
            chat_id,
            TopicKey.CABINET,
            session,
            sender=None,
        )

        if topic_id:
            try:
                try:
                    await _probe_topic_thread(sender, chat_id, topic_id)
                except TelegramBadRequest as e:
                    if "message thread not found" in str(e).lower():
                        topic_id = None
                    else:
                        raise
            except Exception as exc:
                logger.error(f"Failed to validate cabinet topic: {exc}")
                return None

        if not topic_id:
            spec = _get_topic_spec(TopicKey.CABINET)
            if not spec:
                logger.error("Topic spec missing for CABINET")
                return None
            try:
                topic = await sender.create_forum_topic(chat_id, spec.title)
                topic_id = topic.message_thread_id
                await topic_repo.upsert_topic(
                    chat_id=chat_id,
                    key=TopicKey.CABINET.value,
                    thread_id=topic_id,
                    title=spec.title,
                )
                await session.commit()
                invalidate_topic_cache(chat_id)
            except Exception as exc:
                logger.error(f"Failed to create cabinet topic: {exc}")
                return None

        existing_message_id = await repo.get_or_create_panel_message_id(chat_id, topic_id)
        keyboard = _main_keyboard().as_markup()

        if existing_message_id:
            message_id = await _safe_edit_cabinet_message(
                sender,
                repo,
                chat_id,
                topic_id,
                existing_message_id,
                CABINET_HOME_TEXT,
                keyboard,
            )
            await repo.set_panel_message_id(chat_id, topic_id, message_id)
            await session.commit()
            await _pin_cabinet_message(sender, chat_id, topic_id, message_id)
            return topic_id, message_id

        msg = await sender.send_message(
            chat_id=chat_id,
            message_thread_id=topic_id,
            text=CABINET_HOME_TEXT,
            reply_markup=keyboard,
            parse_mode="HTML",
        )
        await repo.get_or_create_panel_message_id(chat_id, topic_id, msg.message_id)
        await session.commit()
        await _pin_cabinet_message(sender, chat_id, topic_id, msg.message_id)
        return topic_id, msg.message_id


def _main_keyboard() -> InlineKeyboardBuilder:
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="📤 Экспорт клиентов", callback_data="cab:export"),
        InlineKeyboardButton(text="📊 Аналитика", callback_data="cab:analytics"),
    )
    builder.row(
        InlineKeyboardButton(text="🔗 Интеграции", callback_data="cab:integrations"),
        InlineKeyboardButton(text="💳 Тариф", callback_data="cab:tariff"),
    )
    return builder


def _stage_keyboard(prefix: str) -> InlineKeyboardBuilder:
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="\u041b\u0438\u0434\u044b", callback_data=f"{prefix}:new"),
        InlineKeyboardButton(text="\u0412 \u0440\u0430\u0431\u043e\u0442\u0435", callback_data=f"{prefix}:in_progress"),
    )
    builder.row(
        InlineKeyboardButton(text="\u041e\u043f\u043b\u0430\u0447\u0435\u043d\u043e", callback_data=f"{prefix}:paid"),
        InlineKeyboardButton(text="Успех", callback_data=f"{prefix}:success"),
    )
    builder.row(
        InlineKeyboardButton(text="\u041e\u0442\u043a\u043b\u043e\u043d\u0435\u043d\u043e", callback_data=f"{prefix}:rejected"),
    )
    builder.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="cab:back"))
    return builder


def _period_keyboard(prefix: str) -> InlineKeyboardBuilder:
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="Сегодня", callback_data=f"{prefix}:today"),
        InlineKeyboardButton(text="Неделя", callback_data=f"{prefix}:week"),
        InlineKeyboardButton(text="Месяц", callback_data=f"{prefix}:month"),
    )
    builder.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="cab:back"))
    return builder


def _period_dates(period: str) -> tuple[datetime, datetime]:
    now = datetime.now(timezone.utc)
    if period == "today":
        return now.replace(hour=0, minute=0, second=0, microsecond=0), now
    if period == "week":
        return now - timedelta(days=7), now
    return now - timedelta(days=30), now


def build_workbook(leads) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = [
        "ID", "Имя", "Телефон", "Эл. почта", "Источник", "Услуга", "Сумма",
        "Статус", "Менеджер", "Комментарий", "UTM", "Создана", "Закрыта",
    ]
    header_fill = PatternFill("solid", fgColor="1A56DB")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_labels = {
        "new": "\u041b\u0438\u0434",
        "in_progress": "\u0412 \u0440\u0430\u0431\u043e\u0442\u0435",
        "paid": "\u041e\u043f\u043b\u0430\u0447\u0435\u043d\u043e",
        "success": "Успех",
        "rejected": "\u041e\u0442\u043a\u043b\u043e\u043d\u0435\u043d\u043e",
    }

    for row, lead in enumerate(leads, 2):
        ws.cell(row=row, column=1, value=lead.id)
        ws.cell(row=row, column=2, value=lead.name)
        ws.cell(row=row, column=3, value=lead.phone)
        ws.cell(row=row, column=4, value=lead.email or "")
        ws.cell(row=row, column=5, value=lead.source)
        ws.cell(row=row, column=6, value=lead.service or "")
        ws.cell(row=row, column=7, value=float(lead.amount) if lead.amount is not None else "")
        ws.cell(row=row, column=8, value=status_labels.get(lead.status.value, lead.status.value))
        ws.cell(row=row, column=9, value=lead.manager.name if lead.manager else "")
        ws.cell(row=row, column=10, value=lead.comment or "")
        ws.cell(row=row, column=11, value=lead.utm_campaign or "")
        ws.cell(row=row, column=12, value=lead.created_at.strftime("%d.%m.%Y %H:%M") if lead.created_at else "")
        ws.cell(row=row, column=13, value=lead.closed_at.strftime("%d.%m.%Y %H:%M") if lead.closed_at else "")

        if row % 2 == 0:
            fill = PatternFill("solid", fgColor="F0F7FF")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@router.message(Command("cabinet"))
async def cmd_cabinet(message: Message, sender: TelegramSafeSender, tenant=None):
    ctx = await resolve_admin_context(message, tenant, sender)
    if not ctx:
        await sender.send_ephemeral_text(
            chat_id=message.chat.id,
            message_thread_id=message.message_thread_id,
            text="⛔️ Кабинет доступен только CRM-администраторам.",
            ttl_sec=TTL_ERROR_SEC,
        )
        return

    await ensure_cabinet_message(sender, message.chat.id)
    try:
        await sender.delete_message(
            chat_id=message.chat.id,
            message_id=message.message_id,
            thread_id=message.message_thread_id,
        )
    except Exception:
        pass


@router.callback_query(F.data == "cab:back")
async def cab_back(callback: CallbackQuery, sender: TelegramSafeSender, tenant=None):
    ctx = await resolve_admin_context(callback, tenant, sender)
    if not ctx:
        await sender.answer(callback)
        return
    await sender.answer(callback)
    if callback.message:
        await ensure_cabinet_message(sender, callback.message.chat.id)


@router.callback_query(F.data == "cab:export")
async def cab_export_menu(callback: CallbackQuery, sender: TelegramSafeSender, tenant=None):
    ctx = await resolve_admin_context(callback, tenant, sender)
    if not ctx:
        await sender.answer(callback, "⛔️ Нет доступа.", show_alert=True)
        return
    await sender.answer(callback)
    await sender.edit_text(
        chat_id=callback.message.chat.id,
        message_id=callback.message.message_id,
        text="📤 Экспорт заявок.\nВыберите этап.",
        reply_markup=_stage_keyboard("cab:export_stage").as_markup(),
        thread_id=callback.message.message_thread_id,
    )


@router.callback_query(F.data.startswith("cab:export_stage:"))
async def cab_export_period(callback: CallbackQuery, sender: TelegramSafeSender, tenant=None):
    ctx = await resolve_admin_context(callback, tenant, sender)
    if not ctx:
        await sender.answer(callback, "⛔️ Нет доступа.", show_alert=True)
        return
    await sender.answer(callback)
    parsed = safe_parse(callback.data, expected_parts=3, expected_types=(str, str, str))
    if not parsed:
        return
    _, _, stage = parsed
    await sender.edit_message_text(
        chat_id=callback.message.chat.id,
        message_id=callback.message.message_id,
        text="📅 <b>Выберите период.</b>",
        reply_markup=_period_keyboard(f"cab:export_do:{stage}").as_markup(),
        parse_mode="HTML",
        thread_id=callback.message.message_thread_id,
    )


@router.callback_query(F.data.startswith("cab:export_do:"))
async def cab_export_do(callback: CallbackQuery, sender: TelegramSafeSender, tenant=None):
    ctx = await resolve_admin_context(callback, tenant, sender)
    if not ctx:
        await sender.answer(callback, "⛔️ Нет доступа.", show_alert=True)
        return
    if tenant is None:
        await sender.answer(callback, "No tenant context.", show_alert=True)
        return
    tenant_id = tenant.id
    await sender.answer(callback)

    parsed = safe_parse(callback.data, expected_parts=4, expected_types=(str, str, str, str))
    if not parsed:
        return
    _, _, stage, period = parsed
    date_from, date_to = _period_dates(period)
    try:
        status = LeadStatus(stage)
    except ValueError:
        await sender.answer(callback, "⚠️ Некорректный этап.", show_alert=True)
        return

    async with AsyncSessionLocal() as session:
        repo = LeadRepository(session)
        leads, total = await repo.get_list(
            status=status,
            date_from=date_from,
            date_to=date_to,
            per_page=10000,
            tenant_id=tenant_id,
        )

    if not leads:
        await sender.send_ephemeral_text(
            chat_id=callback.message.chat.id,
            message_thread_id=callback.message.message_thread_id,
            text="ℹ️ Нет заявок по выбранному фильтру.",
            ttl_sec=TTL_MENU_SEC,
        )
        await ensure_cabinet_message(sender, callback.message.chat.id)
        return

    workbook_bytes = await asyncio.get_event_loop().run_in_executor(None, build_workbook, leads)
    filename = f"leads_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
    await sender.send_document(
        chat_id=callback.message.chat.id,
        message_thread_id=callback.message.message_thread_id,
        document=BufferedInputFile(workbook_bytes, filename=filename),
        caption=f"📤 Экспорт: {total} заявок.\nФайл: {filename}",
        parse_mode=None,
        ttl_sec=TTL_MENU_SEC,
    )
    await ensure_cabinet_message(sender, callback.message.chat.id)


@router.callback_query(F.data == "cab:analytics")
async def cab_analytics(callback: CallbackQuery, sender: TelegramSafeSender, tenant=None):
    ctx = await resolve_admin_context(callback, tenant, sender)
    if not ctx:
        await sender.answer(callback, "⛔️ Нет доступа.", show_alert=True)
        return
    await sender.answer(callback)
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="📈 Конверсия", callback_data="cab:analytics:conversion"),
        InlineKeyboardButton(text="👥 Работа с заявками", callback_data="cab:analytics:activity"),
    )
    builder.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="cab:back"))
    await sender.edit_message_text(
        chat_id=callback.message.chat.id,
        message_id=callback.message.message_id,
        text="📊 <b>Аналитика</b>\nВыберите режим:",
        reply_markup=builder.as_markup(),
        parse_mode="HTML",
        thread_id=callback.message.message_thread_id,
    )


@router.callback_query(F.data.startswith("cab:analytics:conversion"))
async def cab_analytics_conversion(callback: CallbackQuery, sender: TelegramSafeSender, tenant=None):
    ctx = await resolve_admin_context(callback, tenant, sender)
    if not ctx:
        await sender.answer(callback, "⛔️ Нет доступа.", show_alert=True)
        return
    await sender.answer(callback)
    await sender.edit_message_text(
        chat_id=callback.message.chat.id,
        message_id=callback.message.message_id,
        text="📈 <b>Конверсия</b>\nВыберите период:",
        reply_markup=_period_keyboard("cab:analytics_conversion").as_markup(),
        parse_mode="HTML",
        thread_id=callback.message.message_thread_id,
    )


@router.callback_query(F.data.startswith("cab:analytics_conversion:"))
async def cab_analytics_conversion_period(callback: CallbackQuery, sender: TelegramSafeSender, tenant=None):
    ctx = await resolve_admin_context(callback, tenant, sender)
    if not ctx:
        await sender.answer(callback, "⛔️ Нет доступа.", show_alert=True)
        return
    if tenant is None:
        await sender.answer(callback, "No tenant context.", show_alert=True)
        return
    tenant_id = tenant.id
    await sender.answer(callback)
    parsed = safe_parse(callback.data, expected_parts=3, expected_types=(str, str, str))
    if not parsed:
        return
    _, _, period = parsed
    date_from, date_to = _period_dates(period)

    async with AsyncSessionLocal() as session:
        repo = LeadRepository(session)
        stats = await repo.get_conversion_stats(date_from=date_from, date_to=date_to, tenant_id=tenant_id)

    s = stats["by_status"]
    total = stats["total"]
    lines = [
        "📈 <b>Конверсия</b>",
        f"За период с {date_from:%d.%m.%y} - {date_to:%d.%m.%y}",
        f"Всего лидов: {total}",
        f"Взято в работу: {s.get('in_progress', 0)} ({_pct(s.get('in_progress', 0), total)})",
        f"\u041e\u043f\u043b\u0430\u0447\u0435\u043d\u043e: {s.get('paid', 0)} ({_pct(s.get('paid', 0), total)})",
        f"Успех: {s.get('success', 0)} ({_pct(s.get('success', 0), total)})",
        f"\u041e\u0442\u043a\u043b\u043e\u043d\u0435\u043d\u043e: {s.get('rejected', 0)} ({_pct(s.get('rejected', 0), total)})",
    ]

    await sender.send_ephemeral_text(
        chat_id=callback.message.chat.id,
        message_thread_id=callback.message.message_thread_id,
        text="\n".join(lines),
        parse_mode="HTML",
        ttl_sec=TTL_MENU_SEC,
    )
    await ensure_cabinet_message(sender, callback.message.chat.id)


@router.callback_query(F.data.startswith("cab:analytics:activity"))
async def cab_analytics_activity(callback: CallbackQuery, sender: TelegramSafeSender, tenant=None):
    ctx = await resolve_admin_context(callback, tenant, sender)
    if not ctx:
        await sender.answer(callback, "⛔️ Нет доступа.", show_alert=True)
        return
    await sender.answer(callback)
    await sender.edit_message_text(
        chat_id=callback.message.chat.id,
        message_id=callback.message.message_id,
        text="👥 <b>Работа с заявками</b>\nВыберите период:",
        reply_markup=_period_keyboard("cab:analytics_activity").as_markup(),
        parse_mode="HTML",
        thread_id=callback.message.message_thread_id,
    )


@router.callback_query(F.data.startswith("cab:analytics_activity:"))
async def cab_analytics_activity_period(callback: CallbackQuery, sender: TelegramSafeSender, tenant=None):
    ctx = await resolve_admin_context(callback, tenant, sender)
    if not ctx:
        await sender.answer(callback, "⛔️ Нет доступа.", show_alert=True)
        return
    await sender.answer(callback)
    parsed = safe_parse(callback.data, expected_parts=3, expected_types=(str, str, str))
    if not parsed:
        return
    _, _, period = parsed
    async with AsyncSessionLocal() as session:
        repo = LeadRepository(session)
        managers = await repo.get_all_managers(tenant_id=tenant.id)

    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="Все", callback_data=f"cab:analytics_activity_run:{period}:all"))
    for manager in managers:
        builder.row(
            InlineKeyboardButton(
                text=manager.name,
                callback_data=f"cab:analytics_activity_run:{period}:{manager.id}",
            )
        )
    builder.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="cab:back"))
    await sender.edit_text(
        chat_id=callback.message.chat.id,
        message_id=callback.message.message_id,
        text="ℹ️ Выберите сотрудника:",
        reply_markup=builder.as_markup(),
        parse_mode="HTML",
        thread_id=callback.message.message_thread_id,
    )


@router.callback_query(F.data.startswith("cab:analytics_activity_run:"))
async def cab_analytics_activity_run(callback: CallbackQuery, sender: TelegramSafeSender, tenant=None):
    ctx = await resolve_admin_context(callback, tenant, sender)
    if not ctx:
        await sender.answer(callback, "⛔️ Нет доступа.", show_alert=True)
        return
    if tenant is None:
        await sender.answer(callback, "No tenant context.", show_alert=True)
        return
    tenant_id = tenant.id
    await sender.answer(callback)

    parsed = safe_parse(callback.data, expected_parts=4, expected_types=(str, str, str, str))
    if not parsed:
        return
    _, _, period, manager_raw = parsed
    date_from, date_to = _period_dates(period)
    manager_id = None if manager_raw == "all" else (int(manager_raw) if manager_raw.isdigit() else None)

    async with AsyncSessionLocal() as session:
        repo = LeadRepository(session)
        stats = await repo.get_activity_stats(
            date_from=date_from,
            date_to=date_to,
            manager_id=manager_id,
            tenant_id=tenant_id,
        )

    total = stats["total"]
    s = stats["by_status"]
    lines = [
        "👥 <b>Работа с заявками</b>",
        f"За период с {date_from:%d.%m.%y} - {date_to:%d.%m.%y}",
        f"Всего лидов: {total}",
        f"Взято в работу: {s.get('in_progress', 0)} ({_pct(s.get('in_progress', 0), total)})",
        f"\u041e\u043f\u043b\u0430\u0447\u0435\u043d\u043e: {s.get('paid', 0)} ({_pct(s.get('paid', 0), total)})",
        f"Успех: {s.get('success', 0)} ({_pct(s.get('success', 0), total)})",
        f"\u041e\u0442\u043a\u043b\u043e\u043d\u0435\u043d\u043e: {s.get('rejected', 0)} ({_pct(s.get('rejected', 0), total)})",
    ]

    await sender.send_ephemeral_text(
        chat_id=callback.message.chat.id,
        message_thread_id=callback.message.message_thread_id,
        text="\n".join(lines),
        parse_mode="HTML",
        ttl_sec=TTL_MENU_SEC,
    )
    await ensure_cabinet_message(sender, callback.message.chat.id)


def _pct(part: int, total: int) -> str:
    if not total:
        return "0%"
    return f"{round(part / total * 100)}%"


@router.callback_query(F.data == "cab:integrations")
async def cab_integrations(callback: CallbackQuery, sender: TelegramSafeSender, tenant=None):
    ctx = await resolve_admin_context(callback, tenant, sender)
    if not ctx:
        await sender.answer(callback, "⛔️ Нет доступа.", show_alert=True)
        return
    await sender.answer(callback)

    raw_domain = (settings.public_domain or "YOUR_DOMAIN").strip()
    domain = raw_domain.replace("https://", "").replace("http://", "").strip("/") or "YOUR_DOMAIN"
    url = f"https://{domain}/api/v1/leads/tilda"
    escaped_url = html_escape(url)
    api_key = tenant.api_key if tenant else None
    masked_key = f"{api_key[:8]}...{api_key[-4:]}" if api_key else "-"
    escaped_key = html_escape(masked_key)
    text = (
        "🔗 <b>Webhook для Tilda</b>\n"
        f"POST {escaped_url}\n"
        f"X-API-Key: {escaped_key}\n"
        "⚠️ Не вставляйте API-ключ в браузерный JS.\n"
        "📋 Ниже отправлен безопасный proxy-сниппет (без секрета в браузере)."
    )
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="📋 Скопировать ссылку", callback_data="cab:copy_webhook"))
    builder.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="cab:back"))
    await sender.edit_message_text(
        chat_id=callback.message.chat.id,
        message_id=callback.message.message_id,
        text=text,
        reply_markup=builder.as_markup(),
        parse_mode="HTML",
        thread_id=callback.message.message_thread_id,
    )
    try:
        snippet = FSInputFile("integrations/tilda_snippet.js")
        await sender.send_document(
            chat_id=callback.message.chat.id,
            message_thread_id=callback.message.message_thread_id,
            document=snippet,
            caption="Proxy-сниппет для Tilda (без client-side ключа)",
            parse_mode=None,
            ttl_sec=TTL_MENU_SEC,
        )
    except Exception:
        await sender.send_ephemeral_text(
            chat_id=callback.message.chat.id,
            message_thread_id=callback.message.message_thread_id,
            text="⛔️ Не удалось отправить файл JS-сниппета. Проверьте integrations/tilda_snippet.js.",
            ttl_sec=TTL_ERROR_SEC,
        )


@router.callback_query(F.data == "cab:copy_webhook")
async def cab_copy_webhook(callback: CallbackQuery, sender: TelegramSafeSender, tenant=None):
    ctx = await resolve_admin_context(callback, tenant, sender)
    if not ctx:
        await sender.answer(callback)
        return
    await sender.answer(callback)
    raw_domain = (settings.public_domain or "YOUR_DOMAIN").strip()
    domain = raw_domain.replace("https://", "").replace("http://", "").strip("/") or "YOUR_DOMAIN"
    url = f"https://{domain}/api/v1/leads/tilda"
    await sender.send_ephemeral_text(
        chat_id=callback.message.chat.id,
        message_thread_id=callback.message.message_thread_id,
        text=url,
        ttl_sec=TTL_MENU_SEC,
    )


@router.callback_query(F.data == "cab:tariff")
async def cab_tariff(callback: CallbackQuery, sender: TelegramSafeSender, tenant=None):
    ctx = await resolve_admin_context(callback, tenant, sender)
    if not ctx:
        await sender.answer(callback, "⛔️ Нет доступа.", show_alert=True)
        return
    await sender.answer(callback)
    await sender.edit_text(
        chat_id=callback.message.chat.id,
        message_id=callback.message.message_id,
        text="💳 Текущий тариф: StartupImpuls\nЗаявок в месяц: без ограничений\nПоддержка: @StartupImpuls",
        reply_markup=InlineKeyboardBuilder().row(
            InlineKeyboardButton(text="⬅️ Назад", callback_data="cab:back")
        ).as_markup(),
        thread_id=callback.message.message_thread_id,
    )
